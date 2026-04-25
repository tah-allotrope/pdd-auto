"""Spreadsheet-driven Vietnam WTE intake and ProjectInput generation."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import structlog
import yaml
from openpyxl import load_workbook

from pdd_agent.ingest.drive import download_blob
from schemas.project_input import ProjectInput

log = structlog.get_logger(__name__)

REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_MAPPING_CONFIG = REPO_ROOT / "configs" / "source_mappings" / "vietnam_wte_projects.yaml"
DEFAULT_SPREADSHEET_CACHE_DIR = REPO_ROOT / "data" / "source_inputs" / "spreadsheets"
DEFAULT_PROJECTS_DIR = REPO_ROOT / "configs" / "projects"
DEFAULT_REPORTS_DIR = REPO_ROOT / "reports"
DEFAULT_PROJECT_YAML = DEFAULT_PROJECTS_DIR / "vietnam_socson_from_sheet.yaml"
DEFAULT_ASSUMPTIONS_YAML = DEFAULT_PROJECTS_DIR / "vietnam_socson_from_sheet.assumptions.yaml"
DEFAULT_PROFILE_JSON = DEFAULT_SPREADSHEET_CACHE_DIR / "vietnam_wte_profile.json"
DEFAULT_SNAPSHOT_JSON = DEFAULT_SPREADSHEET_CACHE_DIR / "vietnam_socson_snapshot.json"
DEFAULT_REPORT_PATH = DEFAULT_REPORTS_DIR / "source-profile-vietnam-wte.md"


@dataclass
class SpreadsheetArtifacts:
    workbook_path: Path
    profile_json_path: Path
    snapshot_json_path: Path
    project_yaml_path: Path
    assumptions_yaml_path: Path
    report_path: Path


def fetch_workbook(
    mapping_config_path: Path | str = DEFAULT_MAPPING_CONFIG,
    cache_dir: Path | str = DEFAULT_SPREADSHEET_CACHE_DIR,
    force: bool = False,
) -> Path:
    """Download the configured workbook into the stable local cache."""
    config = _load_yaml(mapping_config_path)
    source = config["source_file"]
    cache_dir = Path(cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)

    expected_name = source["expected_name"]
    workbook_path = cache_dir / _safe_filename(expected_name)
    if workbook_path.exists() and workbook_path.stat().st_size > 0 and not force:
        log.info("spreadsheet_cache_hit", path=str(workbook_path))
        return workbook_path

    download_blob(
        source["drive_file_id"],
        workbook_path,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    log.info("spreadsheet_downloaded", path=str(workbook_path))
    return workbook_path


def profile_workbook(workbook_path: Path | str, sample_limit: int = 5) -> dict[str, Any]:
    """Return workbook shape metadata and representative sample rows."""
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)

    sheets: dict[str, Any] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sample_rows = []
        max_sample_row = min(sample_limit, sheet.max_row)
        for row in sheet.iter_rows(min_row=1, max_row=max_sample_row, values_only=True):
            sample_rows.append([_normalize_cell(value) for value in row])
        sheets[sheet_name] = {
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "sample_rows": sample_rows,
        }

    return {
        "workbook_name": workbook_path.name,
        "workbook_path": str(workbook_path),
        "sheet_order": workbook.sheetnames,
        "sheets": sheets,
    }


def select_candidate_row(
    workbook_path: Path | str,
    mapping_config_path: Path | str = DEFAULT_MAPPING_CONFIG,
    candidate_key: str = "soc-son",
) -> dict[str, Any]:
    """Select and normalize one candidate row from the configured workbook."""
    config = _load_yaml(mapping_config_path)
    selection = config["selection"]
    candidate = selection["candidates"][candidate_key]
    workbook = load_workbook(Path(workbook_path), data_only=True, read_only=False)
    sheet = workbook[selection["sheet_name"]]
    header_row = selection.get("header_row", 1)

    headers = _headers_for_sheet(sheet, header_row)
    candidate_column = selection["candidate_column"]
    country_column = selection["country_column"]
    methodology_column = selection["methodology_column"]

    for row_index in range(header_row + 1, sheet.max_row + 1):
        row = [
            sheet.cell(row=row_index, column=index).value for index in range(1, len(headers) + 1)
        ]
        raw_values = dict(zip(headers, (_normalize_cell(value) for value in row), strict=False))
        project_name = str(_get_value(raw_values, candidate_column) or "")
        country = str(_get_value(raw_values, country_column) or "")
        methodology = str(_get_value(raw_values, methodology_column) or "")

        if project_name != candidate["project_name"]:
            continue
        if candidate.get("country_contains") and candidate["country_contains"] not in country:
            continue
        if candidate.get("methodology") and methodology != candidate["methodology"]:
            continue

        normalized = _normalize_project_row(raw_values)
        return {
            "candidate_key": candidate_key,
            "sheet_name": selection["sheet_name"],
            "row_index": row_index,
            "header_row": header_row,
            "raw_values": normalized,
        }

    raise ValueError(f"No workbook row matched candidate {candidate_key!r}")


def generate_project_artifacts(
    workbook_path: Path | str,
    mapping_config_path: Path | str = DEFAULT_MAPPING_CONFIG,
    candidate_key: str = "soc-son",
    output_dir: Path | str | None = None,
) -> SpreadsheetArtifacts:
    """Create profile, row snapshot, ProjectInput YAML, assumptions YAML, and a report."""
    workbook_path = Path(workbook_path)
    mapping_config_path = Path(mapping_config_path)
    output_dir = Path(output_dir) if output_dir else None

    profile = profile_workbook(workbook_path)
    snapshot = select_candidate_row(workbook_path, mapping_config_path, candidate_key)
    assumptions_payload = build_assumption_register(snapshot)
    project_payload = build_project_input_payload(snapshot)
    ProjectInput.model_validate(project_payload)

    if output_dir is None:
        project_yaml_path = DEFAULT_PROJECT_YAML
        assumptions_yaml_path = DEFAULT_ASSUMPTIONS_YAML
        profile_json_path = DEFAULT_PROFILE_JSON
        snapshot_json_path = DEFAULT_SNAPSHOT_JSON
        report_path = DEFAULT_REPORT_PATH
    else:
        output_dir.mkdir(parents=True, exist_ok=True)
        project_yaml_path = output_dir / "vietnam_socson_from_sheet.yaml"
        assumptions_yaml_path = output_dir / "vietnam_socson_from_sheet.assumptions.yaml"
        profile_json_path = output_dir / "vietnam_wte_profile.json"
        snapshot_json_path = output_dir / "vietnam_socson_snapshot.json"
        report_path = output_dir / "source-profile-vietnam-wte.md"

    for path in [
        project_yaml_path,
        assumptions_yaml_path,
        profile_json_path,
        snapshot_json_path,
        report_path,
    ]:
        path.parent.mkdir(parents=True, exist_ok=True)

    profile_json_path.write_text(
        json.dumps(profile, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    snapshot_json_path.write_text(
        json.dumps(snapshot, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    project_yaml_path.write_text(yaml.safe_dump(project_payload, sort_keys=False), encoding="utf-8")
    assumptions_yaml_path.write_text(
        yaml.safe_dump(assumptions_payload, sort_keys=False), encoding="utf-8"
    )
    report_path.write_text(
        render_source_profile_report(profile, snapshot, assumptions_payload, mapping_config_path),
        encoding="utf-8",
    )

    return SpreadsheetArtifacts(
        workbook_path=workbook_path,
        profile_json_path=profile_json_path,
        snapshot_json_path=snapshot_json_path,
        project_yaml_path=project_yaml_path,
        assumptions_yaml_path=assumptions_yaml_path,
        report_path=report_path,
    )


def build_project_input_payload(snapshot: dict[str, Any]) -> dict[str, Any]:
    """Map one normalized spreadsheet row to a validated ProjectInput payload."""
    values = snapshot["raw_values"]
    annual_waste_throughput = float(values["treatment_capacity_tpd"]) * 365.0
    energy_generation_raw = _as_float(values.get("generated_electricity_mwh"), default=0.0)
    net_er_raw = _as_float(values.get("estimated_annual_emission_reductions"), default=0.0)
    energy_generation: float = 0.0 if energy_generation_raw is None else energy_generation_raw
    net_er: float = 0.0 if net_er_raw is None else net_er_raw
    crediting = _parse_crediting_period(values.get("crediting_period_term"))
    waste_type = _derive_waste_types(values)

    return {
        "project": {
            "project_name": values["project_name"],
            "project_id_vcs": _extract_vcs_project_id(values.get("reference_url")),
            "proponent_name": "Soc Son Waste-to-Power Project Company (synthetic assumption)",
            "proponent_contact_email": "socson-pdd-draft@example.com",
            "other_entities": ["Synthetic placeholder pending project sponsor confirmation"],
            "ownership": "Synthetic draft assumption: the project company owns and operates the facility and retains carbon credit ownership pending documentary confirmation.",
        },
        "location": {
            "country": values.get("country_or_province") or "Vietnam",
            "region": "Hanoi",
            "city": "Soc Son",
            "latitude": 21.261,
            "longitude": 105.847,
            "landfill_latitude": 21.275,
            "landfill_longitude": 105.86,
        },
        "dates": {
            "start_date": crediting["start_date"],
            "crediting_period_start": crediting["start_date"],
            "crediting_period_years": crediting["years"],
            "project_scale_small": False,
        },
        "technology": {
            "methodology_ids": [values.get("vcs_methodology") or "ACM0022"],
            "technology_type": "incineration_with_energy_recovery",
            "waste_type": waste_type,
            "annual_waste_throughput": annual_waste_throughput,
            "installed_capacity_mw": _estimate_installed_capacity_mw(energy_generation),
            "energy_generation_mwh_year": energy_generation,
            "tip_fee_usd_per_tonne": 18.0,
            "landfill_diversion_claim": True,
            "fuel_substitution_claim": False,
        },
        "methodology_applicability": {
            "eligibility_checklist": {
                "project treats municipal solid waste": True,
                "waste would otherwise be landfilled": True,
                "electricity displacement is eligible": True,
            },
            "deviation_from_methodology": None,
        },
        "quantification": {
            "baseline_emissions_tco2e_per_year": net_er + 50000.0,
            "project_emissions_tco2e_per_year": 50000.0,
            "leakage_tco2e_per_year": 0.0,
            "net_emissions_tco2e_per_year": net_er,
            "grid_emission_factor": 0.92,
            "grid_emission_factor_source": "Synthetic assumption using Vietnam national grid factor placeholder pending official citation.",
            "methane_capture_rate": 0.2,
            "methane_generation_factor": 0.06,
            "crediting_period_total_tco2e": round(net_er * crediting["years"], 3),
        },
        "monitoring": {
            "parameters_monitored": [
                {
                    "name": "Waste throughput",
                    "unit": "tonnes/day",
                    "frequency": "daily",
                    "method": "weighbridge",
                    "data_source": "Synthetic assumption based on standard WTE operations monitoring",
                },
                {
                    "name": "Net electricity exported",
                    "unit": "MWh",
                    "frequency": "monthly",
                    "method": "revenue meter",
                    "data_source": "Synthetic assumption based on standard power export settlement data",
                },
            ],
            "monitoring_equipment": ["Calibrated weighbridge", "Revenue-grade electricity meter"],
            "data_management": "Synthetic assumption: plant operations data is recorded digitally, reviewed monthly, and archived for verification.",
        },
        "safeguards": {
            "no_net_harm_statement": "Synthetic assumption: a no-net-harm review is required before external use of this draft.",
            "stakeholder_consultation_completed": False,
            "stakeholder_consultation_date": None,
            "environmental_impact_assessment": False,
            "eia_reference": None,
        },
        "compliance_and_ownership": {
            "no_participation_other_programs": True,
            "no_other_forms_of_credit": True,
            "other_ghg_programs": [],
            "credit_ownership_statement": "Synthetic draft assumption: credit ownership remains with the project proponent unless superseded by executed agreements.",
            "double_counting_risk": False,
        },
        "sustainable_development": {
            "sd_contributions": [
                "Improves municipal waste handling",
                "Reduces landfill methane emissions",
                "Supplies lower-carbon electricity to the grid",
            ],
            "sd_comments": "Generated from the Vietnam WTE spreadsheet with explicit synthetic assumptions for missing project facts.",
        },
    }


def build_assumption_register(snapshot: dict[str, Any]) -> dict[str, Any]:
    """Create a machine-readable assumption register for the mapped project."""
    values = snapshot["raw_values"]
    crediting = _parse_crediting_period(values.get("crediting_period_term"))
    net_er_raw = _as_float(values.get("estimated_annual_emission_reductions"), default=0.0)
    net_er: float = 0.0 if net_er_raw is None else net_er_raw
    entries = [
        _entry(
            "project.project_name",
            values["project_name"],
            "spreadsheet",
            "Projects row selection",
            "high",
        ),
        _entry(
            "project.project_id_vcs",
            _extract_vcs_project_id(values.get("reference_url")),
            "spreadsheet",
            "Derived from Verra registry URL",
            "high",
        ),
        _entry(
            "project.proponent_name",
            "Soc Son Waste-to-Power Project Company (synthetic assumption)",
            "synthetic_assumption",
            "Workbook does not identify legal proponent name",
            "medium",
        ),
        _entry(
            "project.proponent_contact_email",
            "socson-pdd-draft@example.com",
            "synthetic_assumption",
            "Workbook does not include contact details",
            "low",
        ),
        _entry(
            "project.other_entities",
            ["Synthetic placeholder pending project sponsor confirmation"],
            "synthetic_assumption",
            "Workbook omits partner entities",
            "low",
        ),
        _entry(
            "project.ownership",
            "Synthetic draft assumption: the project company owns and operates the facility and retains carbon credit ownership pending documentary confirmation.",
            "synthetic_assumption",
            "Workbook omits ownership wording",
            "low",
        ),
        _entry(
            "location.country",
            values.get("country_or_province") or "Vietnam",
            "spreadsheet",
            "Province/Country column",
            "high",
        ),
        _entry(
            "location.region",
            "Hanoi",
            "demo_default",
            "Soc Son is located in Hanoi in the existing repo demo path",
            "medium",
        ),
        _entry(
            "location.city",
            "Soc Son",
            "demo_default",
            "Soc Son alignment with existing demo and local template naming",
            "medium",
        ),
        _entry(
            "location.latitude",
            21.261,
            "synthetic_assumption",
            "Workbook does not contain coordinates",
            "low",
        ),
        _entry(
            "location.longitude",
            105.847,
            "synthetic_assumption",
            "Workbook does not contain coordinates",
            "low",
        ),
        _entry(
            "location.landfill_latitude",
            21.275,
            "synthetic_assumption",
            "Baseline landfill coordinates are missing from the workbook",
            "low",
        ),
        _entry(
            "location.landfill_longitude",
            105.86,
            "synthetic_assumption",
            "Baseline landfill coordinates are missing from the workbook",
            "low",
        ),
        _entry(
            "dates.start_date",
            crediting["start_date"],
            "spreadsheet",
            "Derived from crediting period start in workbook",
            "medium",
        ),
        _entry(
            "dates.crediting_period_start",
            crediting["start_date"],
            "spreadsheet",
            "Crediting Period Term column",
            "medium",
        ),
        _entry(
            "dates.crediting_period_years",
            crediting["years"],
            "spreadsheet",
            "Parsed from crediting period term",
            "high",
        ),
        _entry(
            "technology.methodology_ids",
            [values.get("vcs_methodology") or "ACM0022"],
            "spreadsheet",
            "VCS Methodology column",
            "high",
        ),
        _entry(
            "technology.annual_waste_throughput",
            float(values["treatment_capacity_tpd"]) * 365.0,
            "spreadsheet",
            "Derived from treatment capacity tpd x 365",
            "medium",
        ),
        _entry(
            "technology.installed_capacity_mw",
            _estimate_installed_capacity_mw(_as_float(values.get("generated_electricity_mwh"))),
            "synthetic_assumption",
            "Installed MW is not present, estimated from annual generation using 85% capacity factor",
            "low",
        ),
        _entry(
            "technology.energy_generation_mwh_year",
            _as_float(values.get("generated_electricity_mwh")),
            "spreadsheet",
            "Generated electricity column",
            "high",
        ),
        _entry(
            "technology.tip_fee_usd_per_tonne",
            18.0,
            "demo_default",
            "Reused conservative demo placeholder until commercial inputs are confirmed",
            "low",
        ),
        _entry(
            "quantification.net_emissions_tco2e_per_year",
            net_er,
            "spreadsheet",
            "Estimated Annual Emission Reductions column",
            "high",
        ),
        _entry(
            "quantification.baseline_emissions_tco2e_per_year",
            net_er + 50000.0,
            "synthetic_assumption",
            "Workbook only provides net annual emission reductions, so baseline split is unresolved",
            "low",
        ),
        _entry(
            "quantification.project_emissions_tco2e_per_year",
            50000.0,
            "synthetic_assumption",
            "Workbook does not provide project emissions breakdown",
            "low",
        ),
        _entry(
            "quantification.grid_emission_factor",
            0.92,
            "demo_default",
            "Conservative placeholder reused from existing Vietnam demo config",
            "low",
        ),
        _entry(
            "quantification.grid_emission_factor_source",
            "Synthetic assumption using Vietnam national grid factor placeholder pending official citation.",
            "synthetic_assumption",
            "Official grid factor citation is not in workbook",
            "low",
        ),
        _entry(
            "quantification.crediting_period_total_tco2e",
            round(net_er * crediting["years"], 3),
            "spreadsheet",
            "Net annual ER multiplied by parsed crediting years",
            "medium",
        ),
        _entry(
            "monitoring.parameters_monitored",
            "Standard weighbridge and export metering placeholders",
            "synthetic_assumption",
            "Workbook does not include monitoring plan details",
            "low",
        ),
        _entry(
            "monitoring.data_management",
            "Synthetic assumption: plant operations data is recorded digitally, reviewed monthly, and archived for verification.",
            "synthetic_assumption",
            "Workbook does not include data-management details",
            "low",
        ),
        _entry(
            "safeguards.no_net_harm_statement",
            "Synthetic assumption: a no-net-harm review is required before external use of this draft.",
            "synthetic_assumption",
            "Workbook does not include safeguards documentation",
            "low",
        ),
        _entry(
            "safeguards.environmental_impact_assessment",
            False,
            "synthetic_assumption",
            "Workbook does not evidence an EIA",
            "low",
        ),
        _entry(
            "compliance_and_ownership.credit_ownership_statement",
            "Synthetic draft assumption: credit ownership remains with the project proponent unless superseded by executed agreements.",
            "synthetic_assumption",
            "Workbook does not contain executed ownership language",
            "low",
        ),
    ]

    blocked_review_paths = [
        "quantification.baseline_emissions_tco2e_per_year",
        "quantification.project_emissions_tco2e_per_year",
        "quantification.grid_emission_factor_source",
        "location.latitude",
        "location.longitude",
        "safeguards.no_net_harm_statement",
        "monitoring.parameters_monitored",
    ]

    return {
        "candidate_key": snapshot["candidate_key"],
        "source_snapshot": {
            "sheet_name": snapshot["sheet_name"],
            "row_index": snapshot["row_index"],
        },
        "assumptions": entries,
        "guardrails": {
            "blocked_review_paths": blocked_review_paths,
            "blocked_review_items": [
                {
                    "field_path": path,
                    "reason": "Synthetic or demo-derived value requires human review before PDD use",
                }
                for path in blocked_review_paths
            ],
            "notes": [
                "Synthetic assumptions must not be used to justify methodology deviations.",
                "Net annual ER can be carried from the workbook, but baseline/project split remains review-gated.",
                "This mapping keeps landfill diversion only and blocks any hidden fuel substitution inference.",
            ],
        },
    }


def render_source_profile_report(
    profile: dict[str, Any],
    snapshot: dict[str, Any],
    assumptions_payload: dict[str, Any],
    mapping_config_path: Path,
) -> str:
    """Render a human-readable workbook profile report."""
    lines = [
        "# Vietnam WTE Spreadsheet Source Profile",
        "",
        f"- Workbook: `{profile['workbook_name']}`",
        f"- Mapping config: `{mapping_config_path}`",
        f"- Candidate key: `{snapshot['candidate_key']}`",
        f"- Selected sheet/row: `{snapshot['sheet_name']}` / `{snapshot['row_index']}`",
        "",
        "## Workbook Tabs",
        "",
    ]
    for sheet_name in profile["sheet_order"]:
        sheet = profile["sheets"][sheet_name]
        lines.append(
            f"- `{sheet_name}` rows=`{sheet['max_row']}` cols=`{sheet['max_column']}` sample_rows=`{len(sheet['sample_rows'])}`"
        )

    lines.extend(
        [
            "",
            "## Selected Candidate",
            "",
        ]
    )
    for key, value in snapshot["raw_values"].items():
        lines.append(f"- `{key}`: `{value}`")

    lines.extend(
        [
            "",
            "## Review-Gated Assumptions",
            "",
        ]
    )
    for item in assumptions_payload["guardrails"]["blocked_review_items"]:
        lines.append(f"- `{item['field_path']}`: {item['reason']}")

    return "\n".join(lines) + "\n"


def _load_yaml(path: Path | str) -> dict[str, Any]:
    with open(path, encoding="utf-8") as handle:
        return yaml.safe_load(handle)


def _safe_filename(name: str) -> str:
    stem = Path(name).stem
    suffix = Path(name).suffix or ".xlsx"
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("_")
    return f"{safe_stem}{suffix}"


def _get_value(raw_values: dict[str, Any], key: str) -> Any:
    if key in raw_values:
        return raw_values[key]
    normalized_key = _slugify(key)
    return raw_values.get(normalized_key)


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _headers_for_sheet(sheet, header_row: int) -> list[str]:
    headers: list[str] = []
    for cell in sheet[header_row]:
        raw = _normalize_cell(cell.value)
        headers.append(_slugify(raw) if raw is not None else "project_name")

    seen: dict[str, int] = {}
    deduped: list[str] = []
    for index, header in enumerate(headers):
        candidate = header or f"column_{index + 1}"
        count = seen.get(candidate, 0)
        seen[candidate] = count + 1
        deduped.append(candidate if count == 0 else f"{candidate}_{count + 1}")
    if deduped:
        deduped[0] = "project_name"
    return deduped


def _slugify(value: Any) -> str:
    text = str(value or "").replace("\n", " ").replace("/", " ")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_").lower()
    aliases = {
        "treatment_capacity_tpd": "treatment_capacity_tpd",
        "generated_electricity_mwh": "generated_electricity_mwh",
        "estimated_annual_emission_reductions": "estimated_annual_emission_reductions",
        "crediting_period_term": "crediting_period_term",
        "vcs_methodology": "vcs_methodology",
        "ref": "reference_url",
        "province_country": "country_or_province",
    }
    return aliases.get(text, text)


def _normalize_project_row(raw_values: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(raw_values)
    composition = {
        "wood_pct_wet": raw_values.get("wood_and_wood_products_wet"),
        "paper_pct_wet": raw_values.get("pulp_paper_and_cardboard_wet"),
        "food_pct_wet": raw_values.get("food_food_waste_beverage_and_tobacco_wet"),
        "textiles_pct_wet": raw_values.get("textiles_wet"),
        "garden_pct_wet": raw_values.get("garden_yard_and_park_waste_wet"),
        "glass_pct_wet": raw_values.get("glass_wet"),
        "metal_pct_wet": raw_values.get("metal_wet"),
        "plastics_pct_wet": raw_values.get("plastics_wet"),
        "rubber_pct_wet": raw_values.get("rubber_wet"),
        "other_inert_pct_wet": raw_values.get("other_inert_waste_wet"),
    }
    normalized["waste_composition_pct"] = composition
    return normalized


def _derive_waste_types(values: dict[str, Any]) -> list[str]:
    types = ["municipal_solid_waste"]
    food_share_raw = _as_float(values.get("food_food_waste_beverage_and_tobacco_wet"), default=0.0)
    plastics_share_raw = _as_float(values.get("plastics_wet"), default=0.0)
    food_share: float = 0.0 if food_share_raw is None else food_share_raw
    plastics_share: float = 0.0 if plastics_share_raw is None else plastics_share_raw
    if food_share > 0:
        types.append("food_waste")
    if plastics_share > 0:
        types.append("plastics")
    return types


def _parse_crediting_period(term: Any) -> dict[str, Any]:
    text = str(term or "").strip()
    match = re.search(r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})", text)
    if not match:
        return {"start_date": "2022-07-24", "years": 7}
    start = match.group(1)
    end = match.group(2)
    start_parts = start.split("/")
    end_parts = end.split("/")
    start_iso = f"{start_parts[2]}-{start_parts[1]}-{start_parts[0]}"
    years = int(end_parts[2]) - int(start_parts[2])
    return {"start_date": start_iso, "years": max(years, 1)}


def _extract_vcs_project_id(reference_url: Any) -> str | None:
    text = str(reference_url or "")
    match = re.search(r"/VCS/(\d+)", text)
    if not match:
        return None
    return f"VCS-{match.group(1)}"


def _estimate_installed_capacity_mw(energy_generation_mwh_year: float | None) -> float:
    if not energy_generation_mwh_year:
        return 0.0
    return round(energy_generation_mwh_year / (8760.0 * 0.85), 3)


def _as_float(value: Any, default: float | None = None) -> float | None:
    if value in (None, "", "N/A"):
        return default
    return float(value)


def _entry(
    field_path: str,
    value: Any,
    source_type: str,
    rationale: str,
    confidence: str,
) -> dict[str, Any]:
    return {
        "field_path": field_path,
        "value": value,
        "source_type": source_type,
        "rationale": rationale,
        "confidence": confidence,
    }
