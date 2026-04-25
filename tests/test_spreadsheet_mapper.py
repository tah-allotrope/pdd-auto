"""Tests for Vietnam WTE spreadsheet profiling and ProjectInput mapping."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

ROOT_DIR = Path(__file__).parent.parent.resolve()
sys.path.insert(0, str(ROOT_DIR))
sys.path.insert(0, str(ROOT_DIR / "src"))

from pdd_agent.phase06.spreadsheet_mapper import (
    generate_project_artifacts,
    profile_workbook,
    select_candidate_row,
)
from schemas.project_input import ProjectInput


def _make_workbook(path: Path) -> Path:
    workbook = Workbook()
    projects = workbook.active
    assert projects is not None
    projects.title = "Projects"
    projects.append(
        [
            None,
            "Treatment capacity (tpd)",
            "Generated electricity\nMWh",
            "Estimated Annual Emission Reductions",
            "Wood and wood products\n(%, wet)",
            "Pulp, paper and cardboard\n(%, wet)",
            "Food, food waste, beverage and tobacco\n(%, wet)",
            "Textiles\n(%, wet)",
            "Garden, yard and park waste\n(%, wet)",
            "Glass\n(%, wet)",
            "Metal\n(%, wet)",
            "Plastics\n(%, wet)",
            "Rubber\n(%, wet)",
            "Other, inert waste\n(%, wet)",
            "Crediting Period Term",
            "VCS Methodology",
            "Ref",
            "Province/Country",
        ]
    )
    projects.append(
        [
            "Soc Son waste to power plant project",
            4000.0,
            388050.0,
            544076.0,
            0.0,
            2.7,
            51.9,
            1.6,
            0.0,
            0.5,
            0.9,
            3.0,
            1.3,
            38.1,
            "1st, 24/07/2022 - 23/07/2029",
            "ACM0022",
            "https://registry.verra.org/app/projectDetail/VCS/2567",
            "Vietnam",
        ]
    )

    model = workbook.create_sheet("Model")
    assert model is not None
    model.append(["KEY ASSUMPTIONS", None])
    model.append(["Assumed Avg. Carbon Price ($/ton)", 6.0])

    doxaco = workbook.create_sheet("DOXACO")
    assert doxaco is not None
    doxaco.append([1.0, "Baseline Emissions"])

    claude_log = workbook.create_sheet("Claude Log")
    assert claude_log is not None
    claude_log.append(["Turn #", "Date", "User Request"])

    workbook.save(path)
    return path


def _write_mapping(path: Path) -> Path:
    payload = {
        "source_file": {
            "drive_file_id": "1tMcKxUGE5aIs-3BQ7sJtjKHeOdSLUhKG",
            "expected_name": "WtE plants carbon model early draft.xlsx",
        },
        "selection": {
            "sheet_name": "Projects",
            "header_row": 1,
            "candidate_column": "project_name",
            "country_column": "Province/Country",
            "methodology_column": "VCS Methodology",
            "candidates": {
                "soc-son": {
                    "project_name": "Soc Son waste to power plant project",
                    "country_contains": "Vietnam",
                    "methodology": "ACM0022",
                }
            },
        },
        "profile_tabs": {
            "Model": "assumptions",
            "Projects": "facts",
            "DOXACO": "audit_notes",
            "Claude Log": "audit_notes",
        },
    }
    path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
    return path


def test_profile_workbook_reports_sheet_shapes(tmp_path: Path):
    workbook_path = _make_workbook(tmp_path / "sample.xlsx")

    profile = profile_workbook(workbook_path)

    assert profile["workbook_name"] == "sample.xlsx"
    assert profile["sheets"]["Projects"]["max_row"] == 2
    assert profile["sheets"]["Projects"]["max_column"] == 18
    assert (
        profile["sheets"]["Projects"]["sample_rows"][1][0] == "Soc Son waste to power plant project"
    )


def test_select_candidate_row_normalizes_soc_son_fields(tmp_path: Path):
    workbook_path = _make_workbook(tmp_path / "sample.xlsx")
    mapping_path = _write_mapping(tmp_path / "mapping.yaml")

    snapshot = select_candidate_row(
        workbook_path=workbook_path,
        mapping_config_path=mapping_path,
        candidate_key="soc-son",
    )

    assert snapshot["candidate_key"] == "soc-son"
    assert snapshot["sheet_name"] == "Projects"
    assert snapshot["row_index"] == 2
    assert snapshot["raw_values"]["project_name"] == "Soc Son waste to power plant project"
    assert snapshot["raw_values"]["treatment_capacity_tpd"] == 4000.0
    assert snapshot["raw_values"]["generated_electricity_mwh"] == 388050.0
    assert snapshot["raw_values"]["country_or_province"] == "Vietnam"


def test_generate_project_artifacts_writes_project_and_assumptions(tmp_path: Path):
    workbook_path = _make_workbook(tmp_path / "sample.xlsx")
    mapping_path = _write_mapping(tmp_path / "mapping.yaml")
    output_dir = tmp_path / "out"

    artifacts = generate_project_artifacts(
        workbook_path=workbook_path,
        mapping_config_path=mapping_path,
        candidate_key="soc-son",
        output_dir=output_dir,
    )

    assert artifacts.project_yaml_path.exists()
    assert artifacts.assumptions_yaml_path.exists()
    assert artifacts.profile_json_path.exists()
    assert artifacts.snapshot_json_path.exists()
    assert artifacts.report_path.exists()

    project_input = ProjectInput.model_validate(
        yaml.safe_load(artifacts.project_yaml_path.read_text(encoding="utf-8"))
    )
    assumptions = yaml.safe_load(artifacts.assumptions_yaml_path.read_text(encoding="utf-8"))

    assert project_input.project.project_name == "Soc Son waste to power plant project"
    assert project_input.location.country == "Vietnam"
    assert project_input.technology.methodology_ids == ["ACM0022"]
    assert project_input.technology.annual_waste_throughput == pytest.approx(4000.0 * 365.0)
    assert project_input.technology.energy_generation_mwh_year == pytest.approx(388050.0)
    assert project_input.quantification.net_emissions_tco2e_per_year == pytest.approx(544076.0)

    by_path = {entry["field_path"]: entry for entry in assumptions["assumptions"]}
    assert by_path["project.project_name"]["source_type"] == "spreadsheet"
    assert by_path["location.region"]["source_type"] == "demo_default"
    assert by_path["project.proponent_contact_email"]["source_type"] == "synthetic_assumption"
    assert assumptions["guardrails"]["blocked_review_items"]
    assert (
        "quantification.baseline_emissions_tco2e_per_year"
        in assumptions["guardrails"]["blocked_review_paths"]
    )
